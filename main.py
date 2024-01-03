import streamlit as st
from pdf2image import convert_from_bytes
import pandas as pd
import docx2txt
import docx
import PyPDF2
import io
from reportlab.pdfgen import canvas
from openpyxl import load_workbook



from utils import (
    caesar_encryption, caesar_decryption, rode_encryption, rode_decryption,
    rot13_encryption, rot13_decryption, encrypt_vigenere_cipher, decrypt_vingenere_cipher
)

st.set_page_config(page_title="Chiper App", page_icon="🧊", layout="centered", initial_sidebar_state="expanded", menu_items={
        'About': "# This is a Chiper App. This is an *extremely* cool app!"
    })

st.title('Cipher App')
st.caption('Made by: :blue[Muhammad Faridan Sutariya]')


cipher_options = ['Caesar Cipher', 'Rode', 'Rot13', 'Vigenere']
selected_cipher = st.selectbox('Select Cipher:', cipher_options)

def process_caesar_cipher(message, option, key):
    if option == 'Encryption':
        return caesar_encryption(message.upper(), key)
    else:
        return caesar_decryption(message.upper(), key)

def process_rode_cipher(message, option):
    if option == 'Encryption':
        return rode_encryption(message.upper())
    else:
        return rode_decryption(message.upper())

def process_rot13_cipher(message, option):
    if option == 'Encryption':
        return rot13_encryption(message.upper())
    else:
        return rot13_decryption(message.upper())

def process_vigenere_cipher(message, option, key):
    if option == 'Encryption':
        return encrypt_vigenere_cipher(message.upper(), key)
    else:
        return decrypt_vingenere_cipher(message.upper(), key)

def is_alpha_string(s):
    return s.isalpha()

output_docx = None

# File upload feature
uploaded_file = st.file_uploader('Upload file (each  text must be separated by a semicolon)', type=['xlsx', 'txt', 'pdf', 'docx'])

if uploaded_file is not None:
    file_extension = uploaded_file.name.split('.')[-1].lower()

    if file_extension == 'xlsx':
        try:
            wb = load_workbook(uploaded_file)
            sheet = wb.active
            messages = []
            for row in sheet.iter_rows():
                for cell in row:
                    messages.append(str(cell.value))
        except Exception as e:
            st.error(f"Error reading Excel file: {e}")
            messages = []
        if not messages:
            st.warning('No data found in the Excel file.')

    elif file_extension == 'txt':
        # Handle TXT file
        file_contents = uploaded_file.getvalue().decode("utf-8")
        messages = file_contents.split(';')
        messages = [message.strip() for message in messages]
    
    elif file_extension == 'pdf':
        file_contents = uploaded_file.read()
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_contents))
        
        num_pages = len(pdf_reader.pages)

        messages = []
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            messages.append(page.extract_text())
    
    elif file_extension == 'docx':
        # Handle DOCX file
        text = docx2txt.process(uploaded_file)
        messages = text.split('\n')

    key = None
    if selected_cipher == 'Caesar Cipher':
        key = st.number_input('Key:', value=1)
    elif selected_cipher == "Vigenere":
        key = st.text_input('Key:', help="Example key: GOKIL or WTFIND")
        if not key:
            st.warning('Key cannot be empty for Vigenere Cipher!')
        elif not key.isalpha():
            st.warning('Key for Vigenere Cipher must contain only alphabetic characters!')

    option = st.radio('Select Option:', ('Encryption', 'Decryption'))

    if st.button('Process All'):
        if selected_cipher == 'Vigenere' and (not key or not key.isalpha()):
            st.warning('Key cannot be empty and must contain only alphabetic characters for Vigenere Cipher!')
        else:
            processed_output = []
            for message in messages:
                if selected_cipher == 'Caesar Cipher':
                    result = process_caesar_cipher(message.replace(',', ''), option, key)
                elif selected_cipher == 'Rode':
                    result = process_rode_cipher(message.replace(',', ''), option)
                elif selected_cipher == 'Rot13':
                    result = process_rot13_cipher(message.replace(',', ''), option)
                else:
                    result = process_vigenere_cipher(message.replace(',', ''), option, key)
                
                processed_output.append((message.replace(',', ''), result))

            for input_msg, output_msg in processed_output:
                st.markdown(f'Input: {input_msg}\nProcessed Output: \n```\n{output_msg}\n```')

            if processed_output:
                output_text = "\n\n".join([f'Input: {input_msg}\nProcessed Output: \n{output_msg}' for input_msg, output_msg in processed_output])

                df_output = pd.DataFrame(processed_output, columns=['Input', 'Processed Output'])


                if file_extension == 'txt':
                    file_type = "text/plain"
                    file_extension = "txt"
                    data = output_msg  
                elif file_extension == 'pdf':
                    file_type = "application/pdf"
                    file_extension = "pdf"
                    output_pdf = io.BytesIO()
                    pdf = canvas.Canvas(output_pdf)
                    y_coordinate = 800 
                    for _, output_msg in processed_output:
                        pdf.drawString(100, y_coordinate, output_msg)
                        y_coordinate -= 20 
                    pdf.save()
                    output_pdf.seek(0)
                    data = output_pdf 
                elif file_extension == 'docx':
                    file_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                    file_extension = "docx"
                    output_docx = io.BytesIO()
                    docx_writer = docx.Document()
                    for output in processed_output:
                        docx_writer.add_paragraph(output[1])
                    docx_writer.save(output_docx)
                    output_docx.seek(0)
                    data = output_docx 
                elif file_extension == 'xlsx':
                    # Handle Excel file
                    file_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    file_extension = "xlsx"

                    output_excel = io.BytesIO()
                    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                        df_output['Processed Output'].to_excel(writer, index=False, header=False)  
                    output_excel.seek(0)
                    data = output_excel
                else:
                    st.error('File type not supported or invalid.')

                st.download_button(
                    label="Download Processed Results",
                    data=data,
                    file_name=f"processed_results_{selected_cipher.lower()}.{file_extension}",
                    mime=file_type
                )

# Manual input
else:
    st.subheader(f'{selected_cipher} Cipher')
    option = st.radio('Select Option:', ('Encryption', 'Decryption'))
    message = st.text_input('Enter  Text:')

    if selected_cipher == 'Vigenere':
        key = st.text_input('Key:', help="Example key: GOKIL or WTFIND")
        if not key:
            st.warning('Key cannot be empty for Vigenere Cipher!')
        elif not key.isalpha():
            st.warning('Key for Vigenere Cipher must contain only alphabetic characters!')

        if st.button('Encrypt' if option == 'Encryption' else 'Decrypt') and message and key.isalpha():
            processed_text = process_vigenere_cipher(message, option, key)
            st.markdown(f'Processed Text: \n```\n{processed_text}\n```')
            
    elif selected_cipher == 'Caesar Cipher':
        key = st.number_input('Key:', value=1)
        
        if st.button('Encrypt' if option == 'Encryption' else 'Decrypt'):
            if not message:
                st.warning('Please enter a  Text!')
            elif not message.isalpha():
                 st.warning(' Text must contain only alphabetic characters!')
            else:
                processed_text = process_caesar_cipher(message, option, key)
                st.markdown(f'Processed Text: \n```\n{processed_text}\n```')
    elif selected_cipher in ['Rode', 'Rot13']:
        if st.button('Encrypt' if option == 'Encryption' else 'Decrypt'):
            if not message:
                st.warning('Please enter a  Text!')
            elif not message.isalpha():
                 st.warning(' Text must contain only alphabetic characters!')
            else:
                if selected_cipher == 'Rode':
                    processed_text = process_rode_cipher(message, option)
                else:
                    processed_text = process_rot13_cipher(message, option)

                st.markdown(f'Processed Text: \n```\n{processed_text}\n```')
